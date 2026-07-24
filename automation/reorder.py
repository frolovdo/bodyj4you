"""
BodyJ4You reorder generator (v2).

Reads two inputs:
1. catalog.xlsx - stable reference, lives in project. One row per active SKU.
2. FBA Inventory xlsx - weekly raw export from Amazon Seller Central.

Produces:
- Miami weekly ship list (URGENT/PLANNED/UV/STEEL)
- China weekly ship list
- FBA shipment list (Amazon-ready, kit SKUs for STEEL)
- Monthly China reorder for Bony (4-block xlsx)
- Monthly Miami component demand (HS0002/PB-10/HB0005/JOJOBA)
- Miami component pull list (gauges for this week)

Usage:
    python reorder.py <catalog.xlsx> <fba_inventory.xlsx> <mode> [output_path]
"""
import math
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================================
# Constants
# ============================================================================

URGENT_DAYS_THRESHOLD = 30  # Days < 30 = URGENT
COVER_DAYS = 30  # Weekly ship target: 30 days at FBA before lead time buffer
REORDER_HORIZON_DAYS = 45  # Default monthly horizon

# Canonical lead time by Group, used to fill a blank Lead Time Override only
# when the catalog itself has nothing to learn from (e.g. an export wiped the
# whole column). Group 1 = Miami-produced (7), Group 2 = kits/bundles (10),
# Group 3 = China direct (30). Explicit catalog values always override this.
CANONICAL_LEAD_BY_GROUP = {1: 7, 2: 10, 3: 30}

# Miami section parent SKUs
UV_PARENTS = {"GK0486-Master"}
STEEL_PARENTS = {"GK0541-Master", "GK0715-Master"}

# Components tracked in Miami pull list
PULL_COMPONENTS = [
    "GK0078", "GK0278", "GK0242", "GK0184", "GK0230",
    "GK0279", "GK0220-Gold", "BJGK0183", "GK0315",
]

# Bony China monthly blocks
CHINA_MONTHLY_BLOCK_ORDER = ["GK", "PL6328", "PJ_FJ", "NC"]
CHINA_MONTHLY_BLOCK_LABELS = {
    "GK": "GK (gauges and kits)",
    "PL6328": "PL6328 (taper plugs)",
    "PJ_FJ": "PJ + FJ (piercing and fashion jewelry)",
    "NC": "NC (chokers)",
}

# ============================================================================
# Rounding
# ============================================================================

def roundup_to_10(x):
    if x <= 0:
        return 0
    return int(math.ceil(x / 10.0)) * 10


def roundup_to_100(x):
    if x <= 0:
        return 0
    return int(math.ceil(x / 100.0)) * 100


def roundup_monthly(x):
    """Tiered: >= 500 -> nearest 100, else nearest 10."""
    if x <= 0:
        return 0
    base = math.ceil(x)
    if base >= 500:
        return roundup_to_100(base)
    return roundup_to_10(base)


# ============================================================================
# Loaders
# ============================================================================

def load_catalog(catalog_path):
    """
    Read catalog.xlsx.
    Returns list of dicts in catalog order (preserves the order Denis curated).
    """
    wb = load_workbook(catalog_path, data_only=True)
    ws = wb["Catalog"]

    # Map header name -> column index
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            headers[h] = c

    required = ["SKU", "ASIN", "FBA SKU", "Parent ASIN", "Category",
                "Group", "Warehouse", "China Block", "Lead Time Override",
                "Monthly Days of Cover", "BOM Contributes To"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(f"Catalog missing columns: {missing}")

    # If "Order" column exists use it, otherwise preserve row order
    order_col = headers.get("Order")

    catalog = []
    for r in range(2, ws.max_row + 1):
        asin = ws.cell(r, headers["ASIN"]).value
        if not asin:
            continue
        order_val = ws.cell(r, order_col).value if order_col else (r - 1)
        catalog.append({
            "order": order_val if order_val is not None else (r - 1),
            "sku": ws.cell(r, headers["SKU"]).value,
            "asin": asin,
            "fba_sku": ws.cell(r, headers["FBA SKU"]).value,
            "parent": ws.cell(r, headers["Parent ASIN"]).value,
            "category": ws.cell(r, headers["Category"]).value,
            "group": ws.cell(r, headers["Group"]).value,
            "warehouse": ws.cell(r, headers["Warehouse"]).value,
            "china_block": ws.cell(r, headers["China Block"]).value or "",
            "lead_time": ws.cell(r, headers["Lead Time Override"]).value,
            "monthly_days": ws.cell(r, headers["Monthly Days of Cover"]).value or 45,
            "bom_str": ws.cell(r, headers["BOM Contributes To"]).value or "",
        })
    infer_missing_lead_times(catalog)
    catalog.sort(key=lambda x: x["order"])
    return catalog


def infer_missing_lead_times(catalog):
    """
    Fill any blank Lead Time Override from the pattern the catalog already
    follows, so a SKU with no explicit lead time gets a sensible value instead
    of a flat warehouse default. This runs on every load, so it doesn't matter
    whether an uploaded catalog populated the column — the pipeline always ends
    up with a lead time per SKU.

    Resolution order for a blank lead time:
      1. most common lead time among SKUs sharing its (Group, Category)
      2. most common among its Group
      3. canonical lead by Group (CANONICAL_LEAD_BY_GROUP) — the layer that
         still works when an export wiped the WHOLE column, leaving nothing to
         learn from
      4. warehouse default (China 30, Miami 10)
    Explicit catalog values always win over all of this. Mutates in place.
    """
    from collections import Counter

    def is_blank(v):
        return v in (None, "")

    by_gc = {}
    by_g = {}
    for it in catalog:
        if is_blank(it["lead_time"]):
            continue
        try:
            lt = int(it["lead_time"])
        except (TypeError, ValueError):
            continue
        by_gc.setdefault((it["group"], it["category"]), Counter())[lt] += 1
        by_g.setdefault(it["group"], Counter())[lt] += 1

    def mode(counter):
        return counter.most_common(1)[0][0] if counter else None

    def canon(group):
        try:
            return CANONICAL_LEAD_BY_GROUP.get(int(group))
        except (TypeError, ValueError):
            return None

    for it in catalog:
        if not is_blank(it["lead_time"]):
            continue
        val = mode(by_gc.get((it["group"], it["category"])))
        if val is None:
            val = mode(by_g.get(it["group"]))
        if val is None:
            val = canon(it["group"])
        if val is None:
            val = 30 if it["warehouse"] == "China" else 10
        it["lead_time"] = val
    return catalog


def load_catalog_by_asin(catalog_path):
    """Convenience wrapper: returns dict keyed by ASIN."""
    return {item["asin"]: item for item in load_catalog(catalog_path)}


def parse_bom(bom_str):
    """
    Parse 'HS0002*2; PB-10' style string into list of (base, multiplier).
    """
    if not bom_str:
        return []
    items = []
    for chunk in bom_str.split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "*" in chunk:
            base, mult = chunk.split("*", 1)
            try:
                items.append((base.strip(), int(mult.strip())))
            except ValueError:
                items.append((chunk, 1))
        else:
            items.append((chunk, 1))
    return items


def load_fba_inventory(fba_path):
    """
    Read raw FBA Inventory file (xlsx or csv) - the standard Amazon export.
    Returns dict with TWO keys per record: ASIN and SKU, so callers can
    look up by either. Useful when catalog ASINs are stale but SKUs match.
    """
    import os

    ext = os.path.splitext(fba_path)[1].lower()

    if ext == ".csv":
        rows_iter, headers = _read_csv_fba(fba_path)
    else:
        rows_iter, headers = _read_xlsx_fba(fba_path)

    required = ["asin", "available", "inbound-quantity", "Total Reserved Quantity",
                "days-of-supply", "units-shipped-t7", "units-shipped-t30",
                "units-shipped-t60", "units-shipped-t90", "fba-minimum-inventory-level"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(f"FBA Inventory missing columns: {missing}")

    has_sku = "sku" in headers

    def num(v):
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            try:
                return float(v)
            except ValueError:
                return 0.0
        return 0.0

    by_asin = {}
    by_sku = {}
    for row in rows_iter:
        asin = row[headers["asin"]]
        sku = row[headers["sku"]] if has_sku else None
        if not asin:
            continue
        record = {
            "available": num(row[headers["available"]]),
            "inbound": num(row[headers["inbound-quantity"]]),
            "reserved": num(row[headers["Total Reserved Quantity"]]),
            "days": num(row[headers["days-of-supply"]]),
            "t7": num(row[headers["units-shipped-t7"]]),
            "t30": num(row[headers["units-shipped-t30"]]),
            "t60": num(row[headers["units-shipped-t60"]]),
            "t90": num(row[headers["units-shipped-t90"]]),
            "min_level": num(row[headers["fba-minimum-inventory-level"]]),
        }
        # Aggregate by ASIN
        if asin in by_asin:
            for k in ["available", "inbound", "reserved", "t7", "t30", "t60", "t90", "min_level"]:
                by_asin[asin][k] += record[k]
            by_asin[asin]["days"] = min(by_asin[asin]["days"], record["days"]) if record["days"] > 0 else by_asin[asin]["days"]
        else:
            by_asin[asin] = dict(record)
        # Aggregate by SKU
        if sku:
            if sku in by_sku:
                for k in ["available", "inbound", "reserved", "t7", "t30", "t60", "t90", "min_level"]:
                    by_sku[sku][k] += record[k]
                by_sku[sku]["days"] = min(by_sku[sku]["days"], record["days"]) if record["days"] > 0 else by_sku[sku]["days"]
            else:
                by_sku[sku] = dict(record)

    return {"by_asin": by_asin, "by_sku": by_sku}


def _read_xlsx_fba(fba_path):
    """Read FBA Inventory from xlsx. Returns (row_iter, header_dict)."""
    wb = load_workbook(fba_path, data_only=True)
    sheet_name = "FBA Inventory" if "FBA Inventory" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            headers[h] = c - 1  # Convert to 0-based for consistency with CSV path

    # Build rows as tuples (0-indexed)
    rows = []
    for r in range(2, ws.max_row + 1):
        row = tuple(ws.cell(r, c).value for c in range(1, ws.max_column + 1))
        rows.append(row)
    return rows, headers


def _read_csv_fba(fba_path):
    """Read FBA Inventory from csv. Returns (row_iter, header_dict)."""
    import csv
    rows = []
    headers = {}
    with open(fba_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                headers = {h: idx for idx, h in enumerate(row)}
            else:
                rows.append(tuple(row))
    return rows, headers


def compute_row(cat_item, fba_record):
    """Compute velocity + Status for a SKU."""
    if fba_record is None:
        # ASIN in catalog but NOT in FBA report.
        # This means the listing has been removed/changed/not yet active.
        # Do not trigger revival logic for these.
        return {
            **cat_item,
            "available": 0, "inbound": 0, "reserved": 0,
            "days": 0, "velocity": 0, "min_level": 0,
            "t30": 0,
            "status": False,
            "in_fba": False,
        }
    # Weighted velocity: convert each window to a daily rate, then blend.
    velocity = (fba_record["t7"] / 7 * 0.4 + fba_record["t30"] / 30 * 0.3 +
                fba_record["t60"] / 60 * 0.2 + fba_record["t90"] / 90 * 0.1)
    # Status: True if pipeline >= min level
    pipeline = fba_record["available"] + fba_record["inbound"] + fba_record["reserved"]
    status = pipeline >= fba_record["min_level"]
    return {
        **cat_item,
        "available": fba_record["available"],
        "inbound": fba_record["inbound"],
        "reserved": fba_record["reserved"],
        "days": fba_record["days"],
        "velocity": velocity,
        "min_level": fba_record["min_level"],
        "t30": fba_record["t30"],
        "status": status,
        "in_fba": True,
    }


def load_combined(catalog_path, fba_path):
    """
    Load catalog (ordered list) + FBA, return list of computed rows in catalog order.

    For each catalog item, find FBA record by trying (in order):
    1. ASIN match
    2. FBA SKU match (handles re-listed products where ASIN changed)
    3. Regular SKU match
    """
    catalog_list = load_catalog(catalog_path)
    fba = load_fba_inventory(fba_path)
    by_asin = fba["by_asin"]
    by_sku = fba["by_sku"]

    out = []
    for item in catalog_list:
        record = by_asin.get(item["asin"])
        if record is None and item.get("fba_sku"):
            record = by_sku.get(item["fba_sku"])
        if record is None:
            record = by_sku.get(item["sku"])
        out.append(compute_row(item, record))
    return out


# ============================================================================
# Velocity lock + daily refresh
#
# The weekly FBA file is the ONLY thing that carries sales history (t7/t30/
# t60/t90) — that's what velocity is computed from. Helium10's daily snapshot
# gives us live Available + Inbound but NO sales history. So:
#
#   - When a weekly FBA file is processed, we freeze the demand numbers
#     (velocity, sales-30d, min level, reserved) into a velocity_lock.json,
#     keyed by ASIN.
#   - Each day, a scheduled task pulls fresh Available + Inbound from Helium10
#     and re-runs the report with the FROZEN demand numbers from the lock.
#
# Velocity never recalibrates on its own — only the next FBA upload re-locks
# it. This trades a little velocity freshness for a daily-current view of
# actual inventory movement and days-of-cover.
# ============================================================================

def write_velocity_lock(rows, output_path, snapshot_label=""):
    """
    Freeze per-ASIN demand numbers from a full FBA computation into JSON.
    Written by the weekly flow; read by the daily flow.
    """
    import json

    by_asin = {}
    for r in rows:
        asin = r.get("asin")
        if not asin:
            continue
        by_asin[asin] = {
            "sku": r.get("sku"),
            "velocity": round(float(r.get("velocity") or 0), 4),
            "t30": int(r.get("t30") or 0),
            "min_level": int(r.get("min_level") or 0),
            "reserved": int(r.get("reserved") or 0),
            "days": int(r.get("days") or 0),  # Amazon days-of-supply (reference)
            "in_fba": bool(r.get("in_fba")),
        }
    payload = {"snapshot": snapshot_label, "by_asin": by_asin}
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=0)
    return output_path


def load_fresh_inventory(inventory_path):
    """
    Read a normalized daily inventory snapshot (from Helium10), keyed by ASIN.

    Accepts JSON in either shape:
      {"by_asin": {"B0...": {"available": 10, "inbound": 5, "reserved": 2}}}
      {"B0...": {"available": 10, "inbound": 5}}
    'reserved' is optional per row; when absent the lock's frozen reserved is
    used instead. Sibling SKUs must already be summed into one row per ASIN.
    """
    import json

    with open(inventory_path, "r", encoding="utf-8-sig") as f:
        data = json.load(f)
    if isinstance(data, dict) and "by_asin" in data:
        data = data["by_asin"]

    def num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    out = {}
    for asin, row in (data or {}).items():
        out[asin] = {
            "available": num(row.get("available")),
            "inbound": num(row.get("inbound", row.get("inbound_quantity"))),
            "reserved": (num(row["reserved"]) if "reserved" in row and row["reserved"] is not None else None),
        }
    return out


def load_daily_combined(catalog_path, lock_path, inventory_path):
    """
    Build computed rows for the DAILY refresh: fresh Available + Inbound from
    Helium10, everything else (velocity, sales-30d, min, reserved, Amazon days)
    frozen from the velocity lock. Same row shape as load_combined, so the
    existing build_miami_xlsx / build_china_xlsx work unchanged.
    """
    import json

    catalog_list = load_catalog(catalog_path)
    with open(lock_path, "r", encoding="utf-8-sig") as f:
        lock = json.load(f)
    lock_by_asin = lock.get("by_asin", lock) if isinstance(lock, dict) else {}
    fresh = load_fresh_inventory(inventory_path)

    out = []
    for item in catalog_list:
        asin = item["asin"]
        locked = lock_by_asin.get(asin) or {}
        inv = fresh.get(asin)

        velocity = float(locked.get("velocity") or 0)
        t30 = int(locked.get("t30") or 0)
        min_level = float(locked.get("min_level") or 0)
        amazon_days = float(locked.get("days") or 0)
        locked_reserved = float(locked.get("reserved") or 0)

        if inv is not None:
            available = inv["available"]
            inbound = inv["inbound"]
            # Use fresh reserved if Helium10 provided it, else the frozen value.
            reserved = inv["reserved"] if inv["reserved"] is not None else locked_reserved
            in_fba = True
        else:
            # ASIN not in today's Helium10 snapshot: treat as no shelf stock,
            # keep frozen demand so revival logic can still fire.
            available = 0
            inbound = 0
            reserved = locked_reserved
            in_fba = bool(locked.get("in_fba"))

        pipeline = available + inbound + reserved
        status = pipeline >= min_level
        out.append({
            **item,
            "available": available,
            "inbound": inbound,
            "reserved": reserved,
            "days": amazon_days,
            "velocity": velocity,
            "min_level": min_level,
            "t30": t30,
            "status": status,
            "in_fba": in_fba,
        })
    return out


# ============================================================================
# Calculations
# ============================================================================

def lead_time(row):
    """Lead time: catalog override if present, else 30 for China, 10 for Miami."""
    if row["lead_time"]:
        return float(row["lead_time"])
    if row["warehouse"] == "China":
        return 30
    return 10


def ship_qty(row):
    """
    Weekly ship qty.
    Standard: ROUNDUP10(velocity * (30 + lead_time) - pipeline). Positive only.
    Revival: if Available=0 AND velocity=0, return 20 to wake up the listing.
             This applies whether or not the SKU appears in the FBA report.
             (A catalog SKU not in FBA looks the same as a fully-empty FBA row.)
    """
    if row["available"] == 0 and row["velocity"] == 0:
        return 20

    target = row["velocity"] * (COVER_DAYS + lead_time(row))
    pipeline = row["available"] + row["inbound"] + row["reserved"]
    return roundup_to_10(target - pipeline)


def monthly_qty(row):
    """ROUNDUP(velocity * days_of_cover) with tiered rounding."""
    return roundup_monthly(row["velocity"] * row["monthly_days"])


def display_days_value(row):
    """
    OUR days of coverage — the single source of truth for "days" everywhere in
    the report (classification, sorting, flags, display).

      velocity > 0             -> (Available + Inbound) / velocity
                                  (inbound counts, so 0 on-hand with units
                                  incoming shows real coverage, not a false 0)
      velocity == 0, stock > 0 -> 999  (indefinite coverage, nothing selling)
      velocity == 0, no stock  -> 0

    Amazon's raw days-of-supply is NEVER used here, not even as a fallback. It
    is unreliable near FBA minimum levels and is reference-only.
    """
    if row["velocity"] > 0:
        return round((row["available"] + row["inbound"]) / row["velocity"], 1)
    if row["available"] > 0:
        return 999
    return 0


def is_out_of_stock(row):
    """On-hand only. Independent of Display Days, so an OOS row with inbound
    still reports its true inbound coverage."""
    return row["available"] <= 0


def classify_miami_section(row):
    """URGENT/PLANNED/UV/STEEL for Miami weekly list."""
    if row["parent"] in UV_PARENTS:
        return "UV"
    if row["parent"] in STEEL_PARENTS:
        return "STEEL"
    # URGENT is decided by OUR computed coverage, never Amazon Days.
    if display_days_value(row) < URGENT_DAYS_THRESHOLD:
        return "URGENT"
    return "PLANNED"


def display_sku(row, context):
    """
    miami_reorder / fba_shipment: STEEL uses FBA kit SKU, else regular SKU.
    china_reorder / monthly: always regular SKU.
    """
    if context in ("miami_reorder", "fba_shipment") and row["parent"] in STEEL_PARENTS:
        return row["fba_sku"]
    return row["sku"]


# ============================================================================
# Output formatters
# ============================================================================

HEADER_TSV = "SKU\tASIN\tParent\tCategory\tAvailable\tInbound\tReserved\tDays\tVelocity\tMinLevel\tStatus\tQTY"


def format_tsv_row(row, qty, context):
    return (
        f"{display_sku(row, context)}\t{row['asin']}\t{row['parent']}\t{row['category']}\t"
        f"{int(row['available'])}\t{int(row['inbound'])}\t{int(row['reserved'])}\t"
        f"{display_days_value(row)}\t{row['velocity']:.2f}\t{int(row['min_level'])}\t"
        f"{row['status']}\t{qty}"
    )


def build_miami_text(rows):
    """Miami weekly ship list as text."""
    miami = [r for r in rows if r["warehouse"] == "Miami"]
    sections = {"URGENT": [], "PLANNED": [], "UV": [], "STEEL": []}
    for r in miami:
        q = ship_qty(r)
        if q <= 0:
            continue
        sections[classify_miami_section(r)].append((r, q))

    out = ["# Ship from Miami\n"]
    labels = {"URGENT": "URGENT FBA", "PLANNED": "PLANNED FBA",
              "UV": "UV", "STEEL": "STEEL (kit SKUs)"}
    total = 0
    skus = 0
    for sec in ["URGENT", "PLANNED", "UV", "STEEL"]:
        if not sections[sec]:
            continue
        out.append(f"\n## {labels[sec]} ({len(sections[sec])} SKUs)\n")
        out.append(HEADER_TSV)
        for r, q in sections[sec]:
            out.append(format_tsv_row(r, q, "miami_reorder"))
            total += q
            skus += 1
    out.append(f"\nMiami total: {total:,} units across {skus} SKUs")
    return "\n".join(out)


def build_china_text(rows):
    """China weekly ship list as text."""
    china = [r for r in rows if r["warehouse"] == "China"]
    items = [(r, ship_qty(r)) for r in china]
    items = [(r, q) for r, q in items if q > 0]

    out = [f"# Ship from China\n\n## {len(items)} SKUs\n", HEADER_TSV]
    total = 0
    for r, q in items:
        out.append(format_tsv_row(r, q, "china_reorder"))
        total += q
    out.append(f"\nChina total: {total:,} units across {len(items)} SKUs")
    return "\n".join(out)


def build_china_xlsx(rows, output_path):
    """
    China weekly ship list as xlsx.

    Two-sheet structure, same contract as build_miami_xlsx so the dashboard
    can render it the same way:
      Sheet 1 "China Reorder" - flat dataset, one row per SKU. China has no
                                URGENT/PLANNED/UV/STEEL split, so the "Section"
                                column carries the China Block instead
                                (PL6328 / GK / PJ / NC) for grouping.
      Sheet 2 "Summary"        - block totals + grand total for stat cards.

    Every value the dashboard renders lives in a cell. No recomputation.
    """
    from openpyxl.utils import get_column_letter

    china = [r for r in rows if r["warehouse"] == "China"]
    items = []  # (block, row, qty)
    for r in china:
        q = ship_qty(r)
        if q <= 0:
            continue
        block = r.get("china_block") or "OTHER"
        items.append((block, r, q))

    # Block order for display
    block_order = {"PL6328": 0, "GK": 1, "PJ_FJ": 2, "NC": 3, "OTHER": 4}
    block_labels = {
        "PL6328": "PL6328 TAPER PLUGS",
        "GK": "GK GAUGES & KITS",
        "PJ_FJ": "PJ + FJ JEWELRY",
        "NC": "NC CHOKERS",
        "OTHER": "OTHER",
    }
    # Stable sort by block, preserving catalog order within each block
    items.sort(key=lambda x: block_order.get(x[0], 99))

    wb = Workbook()
    ws = wb.active
    ws.title = "China Reorder"

    def display_days(r):
        # Single source of truth — see display_days_value(). Amazon Days is
        # written to its own column as reference only, never used here.
        return display_days_value(r), is_out_of_stock(r)

    headers = [
        "Section", "SKU", "FBA SKU", "ASIN", "Parent ASIN", "Category",
        "Available", "Inbound", "Reserved",
        "Amazon Days", "Display Days", "Out Of Stock",
        "Weighted Velocity", "Sales 30 Days", "Min Level", "Status", "QTY",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    center = Alignment(horizontal="center")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    block_totals = {}
    block_counts = {}

    row_n = 2
    for block, r, q in items:
        days_val, is_oos = display_days(r)
        label = block_labels.get(block, block)
        ws.cell(row_n, 1, label)
        ws.cell(row_n, 2, display_sku(r, "china_reorder"))
        ws.cell(row_n, 3, r.get("fba_sku") or display_sku(r, "china_reorder"))
        ws.cell(row_n, 4, r["asin"])
        ws.cell(row_n, 5, r["parent"] or "")
        ws.cell(row_n, 6, r["category"] or "")
        ws.cell(row_n, 7, int(r["available"]))
        ws.cell(row_n, 8, int(r["inbound"]))
        ws.cell(row_n, 9, int(r["reserved"]))
        ws.cell(row_n, 10, int(r["days"]))
        ws.cell(row_n, 11, days_val)
        ws.cell(row_n, 12, is_oos)
        ws.cell(row_n, 13, round(r["velocity"], 2))
        ws.cell(row_n, 14, int(r["t30"]))
        ws.cell(row_n, 15, int(r["min_level"]))
        ws.cell(row_n, 16, r["status"])
        ws.cell(row_n, 17, q)
        row_n += 1
        block_totals[label] = block_totals.get(label, 0) + q
        block_counts[label] = block_counts.get(label, 0) + 1

    widths = [22, 22, 22, 13, 19, 9, 10, 10, 10, 11, 12, 12, 16, 13, 10, 8, 8]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    for c, h in enumerate(["Section", "SKU Count", "Total Units"], 1):
        cell = ws2.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ordered_labels = [block_labels[b] for b in
                      ["PL6328", "GK", "PJ_FJ", "NC", "OTHER"]
                      if block_labels[b] in block_counts]
    grand_total = 0
    grand_skus = 0
    row_i = 2
    for label in ordered_labels:
        ws2.cell(row_i, 1, label)
        ws2.cell(row_i, 2, block_counts[label])
        ws2.cell(row_i, 3, block_totals[label])
        grand_total += block_totals[label]
        grand_skus += block_counts[label]
        row_i += 1
    ws2.cell(row_i, 1, "GRAND TOTAL")
    ws2.cell(row_i, 2, grand_skus)
    ws2.cell(row_i, 3, grand_total)

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14

    wb.save(output_path)

    summary = ["# China Weekly Reorder\n"]
    for label in ordered_labels:
        summary.append(f"  {label}: {block_counts[label]} SKUs, {block_totals[label]:,} units")
    summary.append(f"\n  Grand total: {grand_skus} SKUs, {grand_total:,} units")
    summary.append(f"\n  File: {output_path}")
    return output_path, "\n".join(summary)


def build_monthly_text(rows):
    """Monthly reorder forecast for all SKUs."""
    items = [(r, monthly_qty(r)) for r in rows]
    items = [(r, q) for r, q in items if q > 0]

    out = [f"# Monthly Reorder Forecast\n\n## {len(items)} SKUs\n", HEADER_TSV]
    total = 0
    for r, q in items:
        out.append(format_tsv_row(r, q, "monthly"))
        total += q
    out.append(f"\nMonthly total: {total:,} units across {len(items)} SKUs")
    return "\n".join(out)


def build_fba_shipment_text(rows):
    """Amazon FBA shipment list: SKU + QTY only, STEEL uses kit SKU."""
    miami = [r for r in rows if r["warehouse"] == "Miami"]
    sections = {"URGENT": [], "PLANNED": [], "UV": [], "STEEL": []}
    for r in miami:
        q = ship_qty(r)
        if q <= 0:
            continue
        sections[classify_miami_section(r)].append((r, q))

    out = ["# Amazon FBA Shipment List\n",
           "Paste into Amazon shipment creation after Miami confirms.\n"]
    labels = {"URGENT": "URGENT FBA", "PLANNED": "PLANNED FBA",
              "UV": "UV", "STEEL": "STEEL"}
    total = 0
    skus = 0
    for sec in ["URGENT", "PLANNED", "UV", "STEEL"]:
        if not sections[sec]:
            continue
        out.append(f"\n## {labels[sec]} ({len(sections[sec])} SKUs)\n")
        out.append("SKU\tQTY")
        for r, q in sections[sec]:
            out.append(f"{display_sku(r, 'fba_shipment')}\t{q}")
            total += q
            skus += 1
    out.append(f"\nFBA shipment total: {total:,} units across {skus} SKUs")
    return "\n".join(out)


def build_factory_text(rows):
    """
    Miami component demand from monthly forecast through BOM.
    For each base component (HS0002, PB-10, HB0005, JOJOBA), sum consuming SKUs.
    """
    # Group BOM by base SKU
    base_demand = {}  # base -> list of (consumed_sku, monthly_qty, multiplier)
    for r in rows:
        bom = parse_bom(r["bom_str"])
        if not bom:
            continue
        mq = monthly_qty(r)
        if mq <= 0:
            continue
        for base, mult in bom:
            base_demand.setdefault(base, []).append((r["sku"], mq, mult))

    out = ["# Miami Component Demand (45-day forecast)\n",
           "Base components rolled up from monthly forecast through BOM.\n"]

    if not base_demand:
        out.append("\nNo BOM contributions found.")
        return "\n".join(out)

    out.append(f"\n## Summary\n")
    out.append("Base SKU\tTotal Units")

    breakdown_blocks = []
    for base in sorted(base_demand.keys()):
        items = base_demand[base]
        total = sum(qty * mult for _, qty, mult in items)
        out.append(f"{base}\t{total:,}")
        block = [f"\n## {base} = {total:,} units"]
        for sku, qty, mult in items:
            if mult == 1:
                block.append(f"  {sku}: {qty}")
            else:
                block.append(f"  {sku}: {qty} \u00d7 {mult} = {qty * mult}")
        breakdown_blocks.append("\n".join(block))

    out.append("\n---")
    out.append("\n## Breakdown\n")
    out.extend(breakdown_blocks)
    return "\n".join(out)


def build_pull_text(rows):
    """Miami component pull list for this week's ship list."""
    miami = [r for r in rows if r["warehouse"] == "Miami"]
    # Sum ship qty by SKU
    sku_ship = {}
    for r in miami:
        q = ship_qty(r)
        if q > 0:
            sku_ship[r["sku"]] = sku_ship.get(r["sku"], 0) + q

    out = ["# Miami Component Pull List\n",
           "Gauges and components to pull from inventory for this week's ship list.\n",
           "\n## Components\n", "Component SKU\tUnits to Pull"]
    total = 0
    for comp in PULL_COMPONENTS:
        q = sku_ship.get(comp, 0)
        out.append(f"{comp}\t{q}")
        total += q
    out.append(f"\nTotal: {total:,} units across {len(PULL_COMPONENTS)} components")
    return "\n".join(out)


def build_china_monthly_xlsx(rows, output_path):
    """
    Bony China monthly reorder xlsx, 4-block template.
    Sums duplicate SKUs across kit contexts (STEEL gauges).
    """
    # Aggregate by SKU within each block
    sku_aggregate = {}  # sku -> dict
    for r in rows:
        if r["warehouse"] != "China" and r["group"] != 2:
            # Group 2 (bundles) gauge SKUs may have warehouse=Miami but
            # are produced in China. Include them.
            continue
        block = r["china_block"]
        if not block:
            # Group 2 items don't have china_block set. Default to GK for them.
            if r["group"] == 2:
                block = "GK"
            else:
                continue
        q = monthly_qty(r)
        if q <= 0:
            continue
        days = r["monthly_days"]
        sku = r["sku"]
        if sku in sku_aggregate:
            sku_aggregate[sku]["qty"] += q
            sku_aggregate[sku]["days"] = max(sku_aggregate[sku]["days"], days)
            sku_aggregate[sku]["sources"].append(r["parent"])
        else:
            sku_aggregate[sku] = {
                "sku": sku, "qty": q, "days": days,
                "parent": r["parent"], "block": block,
                "sources": [r["parent"]],
            }

    blocks = {k: [] for k in CHINA_MONTHLY_BLOCK_ORDER}
    for item in sku_aggregate.values():
        if item["block"] in blocks:
            blocks[item["block"]].append(item)

    for items in blocks.values():
        # Preserve catalog order (do not re-sort)
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Reorder"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    center = Alignment(horizontal="center")

    block_columns = {"GK": (1, 2, 3), "PL6328": (5, 6, 7),
                     "PJ_FJ": (9, 10, 11), "NC": (13, 14, 15)}

    for block_key, (sc, qc, dc) in block_columns.items():
        ws.cell(1, sc, "SKU").font = header_font
        ws.cell(1, qc, "REORDER").font = header_font
        ws.cell(1, dc, "DAYS OF COVER").font = header_font
        for c in (sc, qc, dc):
            ws.cell(1, c).fill = header_fill
            ws.cell(1, c).alignment = center
        for i, item in enumerate(blocks[block_key], start=2):
            ws.cell(i, sc, item["sku"])
            ws.cell(i, qc, item["qty"]).alignment = center
            ws.cell(i, dc, item["days"]).alignment = center

    for c, w in [(1,22),(2,12),(3,16),(4,2),(5,22),(6,12),(7,16),(8,2),
                 (9,22),(10,12),(11,16),(12,2),(13,22),(14,12),(15,16)]:
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w

    wb.save(output_path)

    summary = ["# China Monthly Reorder for Bony\n"]
    grand_total = 0
    grand_count = 0
    for k in CHINA_MONTHLY_BLOCK_ORDER:
        items = blocks[k]
        t = sum(i["qty"] for i in items)
        grand_total += t
        grand_count += len(items)
        summary.append(f"  {CHINA_MONTHLY_BLOCK_LABELS[k]}: {len(items)} SKUs, {t:,} units")
    summary.append(f"\n  Grand total: {grand_count} SKUs, {grand_total:,} units")

    summed = [i for i in sku_aggregate.values() if len(i["sources"]) > 1]
    if summed:
        summary.append(f"\n  SKUs summed across kit contexts ({len(summed)}):")
        for item in summed:
            summary.append(f"    {item['sku']}: {item['qty']} (from {len(item['sources'])} contexts)")

    summary.append(f"\n  File: {output_path}")
    return output_path, "\n".join(summary)


def build_miami_xlsx(rows, output_path):
    """
    Miami weekly ship list as xlsx.

    Two-sheet structure:
      Sheet 1 "Miami Reorder" - flat dataset, one row per SKU with all values
                                the dashboard renders. No header bands. The
                                "Section" column drives URGENT/PLANNED/UV/STEEL
                                grouping in the dashboard.
      Sheet 2 "Summary"        - section totals + grand total. Drives the
                                stat cards at the top of the dashboard.

    The dashboard MUST be able to render from these sheets alone, with no
    recomputation. Every visualization value lives in a cell.
    """
    from openpyxl.utils import get_column_letter

    miami = [r for r in rows if r["warehouse"] == "Miami"]
    items = []  # (section, row, qty)
    for r in miami:
        q = ship_qty(r)
        if q <= 0:
            continue
        items.append((classify_miami_section(r), r, q))

    # Preserve catalog order, but stable-sort by section so the dashboard sees
    # URGENT block, then PLANNED, then UV, then STEEL.
    section_order = {"URGENT": 0, "PLANNED": 1, "UV": 2, "STEEL": 3}
    items.sort(key=lambda x: section_order[x[0]])

    wb = Workbook()
    ws = wb.active
    ws.title = "Miami Reorder"

    # Single source of truth — see display_days_value(). Amazon Days is written
    # to its own column as reference only, never used here.
    def display_days(r):
        return display_days_value(r), is_out_of_stock(r)

    headers = [
        "Section", "SKU", "FBA SKU", "ASIN", "Parent ASIN", "Category",
        "Available", "Inbound", "Reserved",
        "Amazon Days", "Display Days", "Out Of Stock",
        "Weighted Velocity", "Sales 30 Days", "Min Level", "Status", "QTY",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    center = Alignment(horizontal="center")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    section_labels = {"URGENT": "URGENT FBA", "PLANNED": "PLANNED FBA",
                      "UV": "UV", "STEEL": "STEEL"}
    section_totals = {"URGENT": 0, "PLANNED": 0, "UV": 0, "STEEL": 0}
    section_counts = {"URGENT": 0, "PLANNED": 0, "UV": 0, "STEEL": 0}

    row_n = 2
    for section, r, q in items:
        days_val, is_oos = display_days(r)
        ws.cell(row_n, 1, section_labels[section])
        ws.cell(row_n, 2, display_sku(r, "miami_reorder"))
        ws.cell(row_n, 3, r["fba_sku"] or r["sku"])
        ws.cell(row_n, 4, r["asin"])
        ws.cell(row_n, 5, r["parent"] or "")
        ws.cell(row_n, 6, r["category"] or "")
        ws.cell(row_n, 7, int(r["available"]))
        ws.cell(row_n, 8, int(r["inbound"]))
        ws.cell(row_n, 9, int(r["reserved"]))
        ws.cell(row_n, 10, int(r["days"]))      # Raw Amazon days
        ws.cell(row_n, 11, days_val)            # Computed/displayed days
        ws.cell(row_n, 12, is_oos)              # Out of stock flag
        ws.cell(row_n, 13, round(r["velocity"], 2))
        ws.cell(row_n, 14, int(r["t30"]))       # Raw 30-day units shipped
        ws.cell(row_n, 15, int(r["min_level"]))
        ws.cell(row_n, 16, r["status"])
        ws.cell(row_n, 17, q)
        row_n += 1
        section_totals[section] += q
        section_counts[section] += 1

    widths = [13, 27, 22, 13, 19, 9, 10, 10, 10, 11, 12, 12, 16, 13, 10, 8, 8]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # Summary sheet for dashboard stat cards
    ws2 = wb.create_sheet("Summary")
    summary_headers = ["Section", "SKU Count", "Total Units"]
    for c, h in enumerate(summary_headers, 1):
        cell = ws2.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    summary_rows = []
    grand_total = 0
    grand_skus = 0
    for sec in ["URGENT", "PLANNED", "UV", "STEEL"]:
        summary_rows.append((section_labels[sec], section_counts[sec], section_totals[sec]))
        grand_total += section_totals[sec]
        grand_skus += section_counts[sec]
    summary_rows.append(("GRAND TOTAL", grand_skus, grand_total))

    for i, (label, count, qty) in enumerate(summary_rows, start=2):
        ws2.cell(i, 1, label)
        ws2.cell(i, 2, count)
        ws2.cell(i, 3, qty)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14

    wb.save(output_path)

    summary = ["# Miami Weekly Reorder\n"]
    for sec in ["URGENT", "PLANNED", "UV", "STEEL"]:
        summary.append(f"  {section_labels[sec]}: {section_counts[sec]} SKUs, {section_totals[sec]:,} units")
    summary.append(f"\n  Grand total: {grand_skus} SKUs, {grand_total:,} units")
    summary.append(f"\n  File: {output_path}")
    return output_path, "\n".join(summary)


# ============================================================================
# Main
# ============================================================================

def main():
    # ---- Daily refresh: fresh Helium10 inventory + frozen velocity lock ----
    # Usage: python reorder.py daily <catalog.xlsx> <lock.json> <inventory.json>
    #                                <miami_out.xlsx> <china_out.xlsx>
    if len(sys.argv) >= 2 and sys.argv[1].lower() == "daily":
        if len(sys.argv) < 7:
            print("Usage: python reorder.py daily <catalog.xlsx> <lock.json> "
                  "<inventory.json> <miami_out.xlsx> <china_out.xlsx>")
            sys.exit(1)
        catalog_path, lock_path, inventory_path = sys.argv[2], sys.argv[3], sys.argv[4]
        miami_out, china_out = sys.argv[5], sys.argv[6]
        rows = load_daily_combined(catalog_path, lock_path, inventory_path)
        _, s_m = build_miami_xlsx(rows, miami_out)
        print(s_m)
        print("\n" + "=" * 80 + "\n")
        _, s_c = build_china_xlsx(rows, china_out)
        print(s_c)
        return

    if len(sys.argv) < 4:
        print("Usage: python reorder.py <catalog.xlsx> <fba_inventory.xlsx> "
              "<miami|miami-xlsx|china|china-xlsx|monthly|fba|factory|pull|china-monthly|"
              "velocity-lock|all> [output_path]")
        sys.exit(1)

    catalog_path = sys.argv[1]
    fba_path = sys.argv[2]
    mode = sys.argv[3].lower()
    output_path = sys.argv[4] if len(sys.argv) > 4 else None

    rows = load_combined(catalog_path, fba_path)

    if mode == "miami":
        print(build_miami_text(rows))
    elif mode == "china":
        print(build_china_text(rows))
    elif mode == "monthly":
        print(build_monthly_text(rows))
    elif mode == "fba":
        print(build_fba_shipment_text(rows))
    elif mode == "factory":
        print(build_factory_text(rows))
    elif mode == "pull":
        print(build_pull_text(rows))
    elif mode == "miami-xlsx":
        out = output_path or "/mnt/user-data/outputs/REORDER_MIAMI.xlsx"
        _, s = build_miami_xlsx(rows, out)
        print(s)
    elif mode == "china-xlsx":
        out = output_path or "/mnt/user-data/outputs/REORDER_CHINA_WEEKLY.xlsx"
        _, s = build_china_xlsx(rows, out)
        print(s)
    elif mode == "china-monthly":
        out = output_path or "/mnt/user-data/outputs/REORDER_CHINA.xlsx"
        _, s = build_china_monthly_xlsx(rows, out)
        print(s)
    elif mode == "velocity-lock":
        out = output_path or "/mnt/user-data/outputs/velocity_lock.json"
        write_velocity_lock(rows, out)
        locked = sum(1 for r in rows if r.get("in_fba"))
        print(f"# Velocity lock written\n  {locked} ASINs with live FBA data, "
              f"{len(rows)} catalog rows total\n  File: {out}")
    elif mode == "all":
        print(build_miami_text(rows))
        print("\n" + "=" * 80 + "\n")
        print(build_china_text(rows))
        print("\n" + "=" * 80 + "\n")
        print(build_monthly_text(rows))
        print("\n" + "=" * 80 + "\n")
        print(build_factory_text(rows))
        print("\n" + "=" * 80 + "\n")
        print(build_pull_text(rows))
    else:
        print(f"Unknown mode: {mode}")
        sys.exit(1)


if __name__ == "__main__":
    main()
